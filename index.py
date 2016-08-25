import MySQLdb
import openpyxl
import os

class Database:

    host = 'localhost'
    user = 'root'
    password = 'nvidia123'
    db = 'nvidia'

    def __init__(self):
        self.connection = MySQLdb.connect(self.host, self.user, self.password, self.db)
        self.cursor = self.connection.cursor()

    def insert(self, query):
        try:
            self.cursor.execute(query)
            self.connection.commit()
            print "success"
        except:
            print 'Conflict'
            self.connection.rollback()
            update = "update machine set machine_ip="+ip+" where mac_address="+mac
            db.query("update machine set machine")





    def query(self, query):
        cursor = self.connection.cursor( MySQLdb.cursors.DictCursor )
        cursor.execute(query)

        return cursor.fetchall()

    def __del__(self):
        self.connection.close()

    def retrieve_data_from_spreadsheet(self):
        wb = openpyxl.load_workbook('test.xlsx')
        sheet = wb.get_sheet_by_name('pyexcel_sheet1')

        maxrows = sheet.max_row+1
        maxcolumns = sheet.max_column+1

        print maxcolumns,maxrows
        # for row in sheet.iter_rows(column_offset=2):
        #     for cell in row:
        #         print cell.value
        msg = ''
        hostname = ''
        for i in range(1,maxrows):
            for j in range(1,maxcolumns):
                if j>=maxcolumns-1:
                    msg+="'"+str(sheet.cell(row=i, column=j).value)+"'"
                elif j == 3:
                    msg+="'"+str(sheet.cell(row=i, column=j).value)+"',"
                    hostname = str(sheet.cell(row=i, column=j).value)
                elif j == 4:
                    mac = str(sheet.cell(row=i, column=j).value)
                    msg+="'"+str(sheet.cell(row=i, column=j).value)+"',"
                elif j == 3:
                    ip = str(sheet.cell(row=i, column=j).value)
                    msg+="'"+str(sheet.cell(row=i, column=j).value)+"',"
                elif j == 7:
                    response = os.system("ping -c 1 " + hostname)
                    if response == 0:
                        msg+="'UP',"
                    else:
                        msg+="'Down',"
                else:
                    msg+="'"+str(sheet.cell(row=i, column=j).value)+"',"
            query = "insert into machine values("+msg+")"
            db.insert(query)




if __name__ == "__main__":

    db = Database()
    mac = ''
    ip = ''
    #CleanUp Operation
    # del_query = "DELETE FROM basic_python_database"
    # db.insert(del_query)

    # Data Insert into the table
    # query = "insert into machines values('10.24.142.7','root','nvidia123')"
    #
    #
    # # db.query(query)
    # db.insert(query)

   # Data retrieved from the table
   #  select_query = "select * from machines"
   #
   #  people = db.query(select_query)
   #
   #  for person in people:
   #      print person['machine_ip']+"  "+person['machine_uname']+"  "+person['machine_passwd']
   #
    db.retrieve_data_from_spreadsheet()